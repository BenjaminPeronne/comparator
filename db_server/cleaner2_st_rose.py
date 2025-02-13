import pandas as pd
from openpyxl import load_workbook

def convertir_xls_en_xlsx(input_xls):
    """ Convertit un fichier .xls en .xlsx et retourne le chemin du fichier converti """
    output_xlsx = input_xls.replace(".xls", ".xlsx")
    
    xls = pd.ExcelFile(input_xls, engine="xlrd")
    
    # Sauvegarde toutes les feuilles en .xlsx
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✅ Fichier converti en .xlsx : {output_xlsx}")
    return output_xlsx

def detecter_ligne_en_tete(df, keywords):
    """ Cherche la ligne contenant une liste de mots-clés (ex: 'famille de produits', 'dénomination des produits') """
    for i in range(10):  # Cherche parmi les 10 premières lignes
        if any(any(str(cell).strip().lower() == kw for kw in keywords) for cell in df.iloc[i].values):
            print(f"✅ En-tête détectée pour {keywords} à la ligne {i + 1}")
            return i
    return None

def nettoyer_fichier_xlsx(file_path, output_excel):
    """ Nettoie un fichier Excel en récupérant les valeurs affichées et en nettoyant les données """

    # 1️⃣ Conversion si le fichier est en .xls
    if file_path.endswith(".xls"):
        file_path = convertir_xls_en_xlsx(file_path)

    # 2️⃣ Charger toutes les feuilles du fichier Excel
    wb = load_workbook(file_path, data_only=True)
    sheet_main = wb.worksheets[0]

    # 3️⃣ Lire toutes les valeurs sous forme de tableau
    data_main = [[cell.value for cell in row] for row in sheet_main.iter_rows()]
    df_main = pd.DataFrame(data_main)

    # 4️⃣ Détecter séparément les lignes des en-têtes pour "FAMILLE DE PRODUITS" et "DÉNOMINATION DES PRODUITS"
    famille_header = detecter_ligne_en_tete(df_main, ["famille de produits"])
    produit_header = detecter_ligne_en_tete(df_main, ["dénomination des produits", "libelles", "libellé"])

    if famille_header is None or produit_header is None:
        raise ValueError("❌ Impossible de détecter 'FAMILLE DE PRODUITS' ou 'DÉNOMINATION DES PRODUITS'.")

    # 5️⃣ Définir la ligne d'en-tête détectée pour structurer le tableau
    header_row = min(famille_header, produit_header)  # Utiliser la première en-tête trouvée
    df_main.columns = df_main.iloc[header_row]  # Définit les en-têtes
    df_main = df_main[header_row + 1:].reset_index(drop=True)  # Supprime les lignes inutiles
    df_main = df_main.dropna(axis=1, how='all')  # Supprime les colonnes vides

    # 🛠 **Corriger les noms de colonnes**
    df_main.columns = df_main.columns.astype(str).str.strip().str.lower()
    print(f"🔍 Colonnes détectées après nettoyage : {list(df_main.columns)}")

    # 6️⃣ **Sélectionner la dernière colonne nommée "gencode"**
    gencode_cols = [col for col in df_main.columns if "gencode" in col]
    if gencode_cols:
        last_gencode_col = gencode_cols[-1]  # Prendre la dernière occurrence
        df_main = df_main.rename(columns={last_gencode_col: "gencode"})  # Renommer en "gencode"
        df_main = df_main.loc[:, ~df_main.columns.duplicated()]  # Supprimer les colonnes dupliquées
        print(f"✅ La dernière colonne 'gencode' a été sélectionnée et renommée.")

    # 7️⃣ Remplir les valeurs de "FAMILLE DE PRODUITS"
    famille_col = next((col for col in df_main.columns if "famille de produits" in col.lower()), None)

    if famille_col:
        df_main[famille_col] = df_main[famille_col].ffill()  # Remplit les cellules fusionnées
        print(f"✅ Colonne '{famille_col}' détectée et appliquée sur toutes les lignes.")
    else:
        print("❌ Colonne 'FAMILLE DE PRODUITS' non trouvée dans le fichier.")

    # 8️⃣ Détection automatique de la colonne produit
    produit_col = next((col for col in ["dénomination des produits", "libelles", "libellé"] if col in df_main.columns), None)

    if produit_col is None:
        raise KeyError(f"❌ Aucune colonne produit trouvée ! Colonnes disponibles : {list(df_main.columns)}")

    print(f"✅ Colonne produit détectée : {produit_col}")

    # 9️⃣ Supprimer les lignes vides dans la colonne produit
    df_cleaned = df_main.dropna(subset=[produit_col])

    # 🔟 Sauvegarde en .xlsx propre
    df_cleaned.to_excel(output_excel, index=False, engine="openpyxl")

    print(f"✅ Nettoyage terminé ! Fichier enregistré sous : 📂 {output_excel}")

# 🔥 **Exécution du script**
file_path = "/Users/benjaminperonne/Documents/Developer/University_Developer/M2_S9_MIAGE/Projet/comparator/db_server/src/BQP SUPER U SAINTE-ROSE  NOV 2019.xls"  # Remplacez par votre fichier
output_excel = "BQP_Nettoye_Gencode1.xlsx"
nettoyer_fichier_xlsx(file_path, output_excel)
